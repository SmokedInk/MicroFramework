# -*- coding: utf-8 -*-
# @Author:  SmokedInk
# @Title:   IOExcel
# @Time:    2019-11-06 12:41:51

import xlrd
import xlwt
import openpyxl
from xlutils.copy import copy


def sheet_style(h, w, sheet, name, height, bold=False):
    """
    设置Excel格式样式
    :param h: 表格行序号
    :param w: 表格列序号
    :param sheet: 表格工作表
    :param name: 字体名称
    :param height: 字体大小
    :param bold: 加粗
    :return: 样式
    """
    style = xlwt.XFStyle()
    al = xlwt.Alignment()
    al.horz = 0x02  # 设置水平居中
    al.vert = 0x01  # 设置垂直居中
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.alignment = al
    style.font = font
    sheet.col(w).width = 100 * 40
    tall_style = xlwt.easyxf('font:height 350')
    first_row = sheet.row(h)
    first_row.set_style(tall_style)
    return style


def write_excel_xls(file_path, data_text, row_headline=None, col_headline=None):
    """
    初次新建Excel-xls文件并写入数据
    :param file_path: 文件路径
    :param data_text: 数据内容
    :param row_headline: 行标题
    :param col_headline: 列标题
    :return: None
    """
    wt = xlwt.Workbook()
    sheet = wt.add_sheet('账户信息表', cell_overwrite_ok=True)
    # 写入行标题
    if row_headline:
        for c in range(len(row_headline)):
            sheet.write(0, c, row_headline[c], sheet_style(0, c, sheet, "Times New Roman", 220))
        rh = 1
    else:
        rh = 0
    # 写入列标题
    if col_headline:
        for r in range(len(col_headline)):
            sheet.write(r, 0, col_headline[r], sheet_style(r, 0, sheet, "Times New Roman", 220))
        ch = 1
    else:
        ch = 0
    # 写入其余数据
    try:
        for i in range(len(data_text)):
            for j in range(0, len(data_text[i])):
                sheet.write(i + rh, j + ch, data_text[i][j], sheet_style(i + rh, j + ch, sheet, 'Times New Roman', 220))
        wt.save(file_path)
        print("写入数据成功")
    except Exception as err:
        print("写入失败", err)


def written_excel_xls(file_path, data_text):
    """
    续写现有的Excel-xls文件
    :param file_path: 文件路径
    :param data_text: 数据内容
    :return: None
    """
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    sheet = wb.sheets()[0]
    col_page = sheet.ncols  # 获取列数
    row_page = sheet.nrows  # 获取行数
    # 复制原有Excel内容
    new_wb = copy(wb)
    sheet = new_wb.get_sheet(0)
    ch = col_page - len(data_text[0])
    # 续写内容
    try:
        for i in range(len(data_text)):
            for j in range(0, len(data_text[i])):
                sheet.write(row_page, j + ch, data_text[i][j],
                              sheet_style(row_page, j + ch, sheet, 'Times New Roman', 220))
        new_wb.save(file_path)
        print("续写成功")
    except Exception as err:
        print("续写失败", err)


def update_excel_xls(file_path, row_num, col_num, data_text):
    """
    修改指定单元格内容
    :param file_path: 文件路径
    :param row_num: 行序号
    :param col_num: 列序号
    :param data_text: 修改文本
    :return: None
    """
    wb = xlrd.open_workbook(file_path)
    # 对数据表格进行复制
    new_wb = copy(wb)
    # 定位到Sheet1表
    sheet = new_wb.get_sheet(0)
    try:
        # 在sheet1表中写入内容
        sheet.write(row_num, col_num, data_text)
        # 对修改后的内容进行保存
        new_wb.save(file_path)
    except IOError as err:
        print("修改失败", err)


def read_excel_xls(file_path):
    """
    读取Excel-xls文件内容
    :param file_path: Excel文件路径
    :return: 文件内容
    """
    workbook = xlrd.open_workbook(file_path)
    # 获取sheet表
    sheet = workbook.sheet_by_index(0)
    # 获取sheet表里的数据（指定行与列）
    sheet_val_01 = sheet.cell_value(1, 1)
    # 获取某一行/列数据（指定行/列）
    sheet_val_02 = sheet.row_values(1)
    sheet_val_03 = sheet.col_values(1)
    # 获取所以数据
    sheet_values = []
    for i in range(sheet.nrows):
        sheet_val = sheet.row_values(i)
        sheet_values.append(sheet_val)
        # print(sheet_val)
    return sheet_values


def write_excel_xlsx(file_path, values):
    """
    初次新建Excel-xlsx文件并写入数据
    :param file_path: 文件路径
    :param values: 数据列表
    :return:
    """
    # 新建xlsx表格对象
    workbook = openpyxl.Workbook()
    # 新建sheet工作簿(index: 索引; title: 工作簿名称)
    worksheet = workbook.create_sheet(index=0, title="sheet1")
    # 写入数据
    for value in values:
        worksheet.append(value)
    # 保存文件
    workbook.save(file_path)


def written_excel_xlsx(file_path, values):
    """
    续写现有的Excel-xlsx文件
    :param file_path: 文件路径
    :param values: 数据列表
    :return: None
    """
    # 读取现有xlsx表格对象
    workbook = openpyxl.load_workbook(file_path)
    # 获取所有的工作簿名称
    sheet_names = workbook.get_sheet_names()
    # 获取工作簿
    worksheet = workbook[sheet_names[0]]
    # 追加数据
    for value in values:
        worksheet.append(value)
    # 保存文件
    workbook.save(file_path)


def update_excel_xlsx(file_path, row_num, col_num, data_text):
    """
    修改指定单元格内容
    :param file_path: 文件路径
    :param row_num: 行序号
    :param col_num: 列序号
    :param data_text: 修改文本
    :return: None
    """
    workbook = openpyxl.load_workbook(file_path)
    # 获取所有的工作簿名称
    sheet_names = workbook.get_sheet_names()
    # 获取工作簿
    worksheet = workbook[sheet_names[0]]
    # 更新数据
    worksheet.cell(row_num, col_num).value = data_text  # 更改已经存在的测试数据
    # 保存文件
    workbook.save(file_path)


def read_excel_xlsx(file_path):
    """
    读取Excel-xls文件内容
    :param file_path: 文件路径
    :return: 文件内容
    """
    # 读取现有xlsx表格对象
    workbook = openpyxl.load_workbook(file_path)
    # 获取所有的工作簿名称
    sheet_names = workbook.get_sheet_names()
    # 获取工作簿
    worksheet = workbook[sheet_names[0]]
    # 获取总行数
    rows = worksheet.max_row
    # 获取总列数
    columns = worksheet.max_column
    # 获取sheet表里的数据（指定行与列）
    sheet_val_01 = worksheet.cell(row=4, column=2).value
    # 获取某一行/列数据（指定行/列）
    sheet_val_02 = [val.value for val in worksheet[1]]      # 指定行
    sheet_val_03 = [val.value for val in worksheet['A']]    # 指定列
    # 获取所有数据
    sheet_values = []
    for i in range(rows):
        sheet_obj = worksheet[i + 1]
        sheet_val = [val.value for val in sheet_obj]
        sheet_values.append(sheet_val)
    # print(sheet_values)
    # 关闭文件
    workbook.close()
    return sheet_values


if __name__ == '__main__':
    file_path = r'..\data\account.xlsx'
    row_headline = ["nickName", "userName", "passWord", "mediaID"]
    data_text = [["宠物汇图", "2493438960", "w1989863.", "17139328"],
                 ["时尚汇图", "3586910388", "annj684627", "17140533"],
                 ["漫漫路远兮", "2478637939", "zxc123456789.", "17140357"]]
    data_text_01 = [["安然雅诗", "3583884076", "hhhh6666", "17278627"]]
    # write_excel(file_path, data_text, row_headline=row_headline)
    # written_excel(file_path, data_text_01)
    # update_excel(file_path, 6, 6, "测试修改")
    text = read_excel_xls(file_path)
    # write_excel_xlsx(file_path, data_text)
    # written_excel_xlsx(file_path, data_text_01)
    # sheet_values = read_excel_xlsx(file_path)
    # print(sheet_values)
